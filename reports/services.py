"""
Servicios para generación de reportes PDF.
"""

from django.db import connection
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
# from weasyprint import HTML, CSS  # Comentado por problemas en Windows
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime, date
from decimal import Decimal

from core.models import Company, Period
from accounting.models_accounts import Account
from accounting.models_journal import JournalEntry, JournalEntryLine


class ReportService:
    """
    Servicio principal para generación de reportes.
    """
    
    def __init__(self, company):
        self.company = company
    
    def generate_balance_sheet(self, period_end, format='pdf'):
        """
        Genera Balance General.
        """
        # Obtener cuentas y saldos
        accounts_data = self._get_balance_sheet_data(period_end)
        
        context = {
            'company': self.company,
            'period_end': period_end,
            'report_date': timezone.now(),
            'accounts_data': accounts_data,
            'title': 'BALANCE GENERAL'
        }
        
        if format == 'pdf':
            return self._generate_pdf_report('reports/balance_sheet.html', context)
        elif format == 'excel':
            return self._generate_excel_balance_sheet(context)
    
    def generate_income_statement(self, period_start, period_end, format='pdf'):
        """
        Genera Estado de Resultados.
        """
        accounts_data = self._get_income_statement_data(period_start, period_end)
        
        context = {
            'company': self.company,
            'period_start': period_start,
            'period_end': period_end,
            'report_date': timezone.now(),
            'accounts_data': accounts_data,
            'title': 'ESTADO DE RESULTADOS'
        }
        
        if format == 'pdf':
            return self._generate_pdf_report('reports/income_statement.html', context)
        elif format == 'excel':
            return self._generate_excel_income_statement(context)
    
    def generate_trial_balance(self, period_end, format='pdf'):
        """
        Genera Balance de Prueba.
        """
        accounts_data = self._get_trial_balance_data(period_end)
        
        context = {
            'company': self.company,
            'period_end': period_end,
            'report_date': timezone.now(),
            'accounts_data': accounts_data,
            'title': 'BALANCE DE PRUEBA'
        }
        
        if format == 'pdf':
            return self._generate_pdf_report('reports/trial_balance.html', context)
        elif format == 'excel':
            return self._generate_excel_trial_balance(context)
    
    def generate_general_ledger(self, account_id, period_start, period_end, format='pdf'):
        """
        Genera Libro Mayor de una cuenta.
        """
        account = Account.objects.get(id=account_id, chart_of_accounts__company=self.company)
        movements = self._get_account_movements(account, period_start, period_end)
        
        context = {
            'company': self.company,
            'account': account,
            'period_start': period_start,
            'period_end': period_end,
            'report_date': timezone.now(),
            'movements': movements,
            'title': f'LIBRO MAYOR - {account.code} {account.name}'
        }
        
        if format == 'pdf':
            return self._generate_pdf_report('reports/general_ledger.html', context)
        elif format == 'excel':
            return self._generate_excel_general_ledger(context)
    
    def generate_aging_report(self, report_date, report_type='receivables', format='pdf'):
        """
        Genera reporte de cartera vencida.
        """
        if report_type == 'receivables':
            aging_data = self._get_receivables_aging(report_date)
            title = 'CARTERA VENCIDA - CUENTAS POR COBRAR'
        else:
            aging_data = self._get_payables_aging(report_date)
            title = 'CARTERA VENCIDA - CUENTAS POR PAGAR'
        
        context = {
            'company': self.company,
            'report_date': report_date,
            'aging_data': aging_data,
            'report_type': report_type,
            'title': title
        }
        
        if format == 'pdf':
            return self._generate_pdf_report('reports/aging_report.html', context)
        elif format == 'excel':
            return self._generate_excel_aging_report(context)
    
    def _get_balance_sheet_data(self, period_end):
        """
        Obtiene datos para Balance General.
        """
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    a.code,
                    a.name,
                    at.type as account_type,
                    at.nature,
                    COALESCE(SUM(
                        CASE 
                            WHEN at.nature = 'debit' THEN jel.debit - jel.credit
                            ELSE jel.credit - jel.debit
                        END
                    ), 0) as balance
                FROM accounting_accounts a
                JOIN accounting_account_types at ON a.account_type_id = at.code
                LEFT JOIN accounting_journal_entry_lines jel ON jel.account_id = a.id
                LEFT JOIN accounting_journal_entries je ON jel.journal_entry_id = je.id
                WHERE a.chart_of_accounts_id = %s
                    AND a.is_detail = true
                    AND (je.date IS NULL OR je.date <= %s)
                    AND (je.status IS NULL OR je.status = 'posted')
                    AND at.is_balance_sheet = true
                GROUP BY a.id, a.code, a.name, at.type, at.nature
                HAVING COALESCE(SUM(
                    CASE 
                        WHEN at.nature = 'debit' THEN jel.debit - jel.credit
                        ELSE jel.credit - jel.debit
                    END
                ), 0) != 0
                ORDER BY a.code
            """, [self.company.chart_of_accounts.id, period_end])
            
            return cursor.fetchall()
    
    def _get_income_statement_data(self, period_start, period_end):
        """
        Obtiene datos para Estado de Resultados.
        """
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    a.code,
                    a.name,
                    at.type as account_type,
                    at.nature,
                    COALESCE(SUM(
                        CASE 
                            WHEN at.nature = 'debit' THEN jel.debit - jel.credit
                            ELSE jel.credit - jel.debit
                        END
                    ), 0) as balance
                FROM accounting_accounts a
                JOIN accounting_account_types at ON a.account_type_id = at.code
                LEFT JOIN accounting_journal_entry_lines jel ON jel.account_id = a.id
                LEFT JOIN accounting_journal_entries je ON jel.journal_entry_id = je.id
                WHERE a.chart_of_accounts_id = %s
                    AND a.is_detail = true
                    AND je.date BETWEEN %s AND %s
                    AND je.status = 'posted'
                    AND at.is_income_statement = true
                GROUP BY a.id, a.code, a.name, at.type, at.nature
                HAVING COALESCE(SUM(
                    CASE 
                        WHEN at.nature = 'debit' THEN jel.debit - jel.credit
                        ELSE jel.credit - jel.debit
                    END
                ), 0) != 0
                ORDER BY a.code
            """, [self.company.chart_of_accounts.id, period_start, period_end])
            
            return cursor.fetchall()
    
    def _get_trial_balance_data(self, period_end):
        """
        Obtiene datos para Balance de Prueba.
        """
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    a.code,
                    a.name,
                    COALESCE(SUM(jel.debit), 0) as total_debit,
                    COALESCE(SUM(jel.credit), 0) as total_credit,
                    COALESCE(SUM(jel.debit - jel.credit), 0) as balance
                FROM accounting_accounts a
                LEFT JOIN accounting_journal_entry_lines jel ON jel.account_id = a.id
                LEFT JOIN accounting_journal_entries je ON jel.journal_entry_id = je.id
                WHERE a.chart_of_accounts_id = %s
                    AND a.is_detail = true
                    AND (je.date IS NULL OR je.date <= %s)
                    AND (je.status IS NULL OR je.status = 'posted')
                GROUP BY a.id, a.code, a.name
                ORDER BY a.code
            """, [self.company.chart_of_accounts.id, period_end])
            
            return cursor.fetchall()
    
    def _get_account_movements(self, account, period_start, period_end):
        """
        Obtiene movimientos de una cuenta específica.
        """
        movements = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__date__range=[period_start, period_end],
            journal_entry__status='posted'
        ).select_related('journal_entry').order_by('journal_entry__date', 'journal_entry__number')
        
        return movements
    
    def _get_receivables_aging(self, report_date):
        """
        Obtiene datos de cartera vencida CxC.
        """
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    c.code,
                    c.business_name,
                    i.number,
                    i.date,
                    i.due_date,
                    i.total_amount,
                    i.balance,
                    CASE 
                        WHEN i.due_date >= %s THEN 'current'
                        WHEN i.due_date >= %s - INTERVAL '30 days' THEN '1-30'
                        WHEN i.due_date >= %s - INTERVAL '60 days' THEN '31-60'
                        WHEN i.due_date >= %s - INTERVAL '90 days' THEN '61-90'
                        ELSE '90+'
                    END as aging_bucket
                FROM ar_invoices i
                JOIN ar_customers c ON i.customer_id = c.id
                WHERE i.company_id = %s
                    AND i.balance > 0
                    AND i.status IN ('sent', 'partial', 'overdue')
                ORDER BY c.business_name, i.due_date
            """, [report_date, report_date, report_date, report_date, self.company.id])
            
            return cursor.fetchall()
    
    def _get_payables_aging(self, report_date):
        """
        Obtiene datos de cartera vencida CxP.
        """
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    s.code,
                    s.business_name,
                    pi.number,
                    pi.date,
                    pi.due_date,
                    pi.total_amount,
                    pi.balance,
                    CASE 
                        WHEN pi.due_date >= %s THEN 'current'
                        WHEN pi.due_date >= %s - INTERVAL '30 days' THEN '1-30'
                        WHEN pi.due_date >= %s - INTERVAL '60 days' THEN '31-60'
                        WHEN pi.due_date >= %s - INTERVAL '90 days' THEN '61-90'
                        ELSE '90+'
                    END as aging_bucket
                FROM ap_purchase_invoices pi
                JOIN ap_suppliers s ON pi.supplier_id = s.id
                WHERE pi.company_id = %s
                    AND pi.balance > 0
                    AND pi.status IN ('received', 'approved', 'partial', 'overdue')
                ORDER BY s.business_name, pi.due_date
            """, [report_date, report_date, report_date, report_date, self.company.id])
            
            return cursor.fetchall()
    
    def _generate_pdf_report(self, template_name, context):
        """
        Genera reporte PDF usando ReportLab.
        """
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Contenido
        story = []
        
        # Encabezado
        story.append(Paragraph(context['company'].name, title_style))
        story.append(Paragraph(f"NIT: {context['company'].tax_id}", styles['Normal']))
        story.append(Paragraph(context['title'], title_style))
        
        if 'period_end' in context:
            story.append(Paragraph(f"Al {context['period_end']}", styles['Normal']))
        elif 'period_start' in context and 'period_end' in context:
            story.append(Paragraph(f"Del {context['period_start']} al {context['period_end']}", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Tabla de datos
        if 'accounts_data' in context:
            # Preparar datos para la tabla
            table_data = [['Código', 'Cuenta', 'Saldo']]
            for account in context['accounts_data']:
                table_data.append([
                    str(account[0]),
                    str(account[1]),
                    f"{float(account[4]):,.0f}"
                ])
            
            # Crear tabla
            table = Table(table_data, colWidths=[1.5*inch, 4*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        # Construir PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return pdf_buffer
    
    def _generate_excel_balance_sheet(self, context):
        """
        Genera Balance General en Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance General"
        
        # Estilos
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        
        # Encabezado
        ws['A1'] = context['company'].name
        ws['A1'].font = header_font
        ws['A2'] = context['title']
        ws['A2'].font = title_font
        ws['A3'] = f"Al {context['period_end']}"
        
        # Cabeceras de columnas
        row = 5
        headers = ['Código', 'Cuenta', 'Saldo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
        
        # Datos
        row += 1
        for account_data in context['accounts_data']:
            ws.cell(row=row, column=1, value=account_data[0])  # código
            ws.cell(row=row, column=2, value=account_data[1])  # nombre
            ws.cell(row=row, column=3, value=float(account_data[4]))  # saldo
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _generate_excel_income_statement(self, context):
        """
        Genera Estado de Resultados en Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estado de Resultados"
        
        # Estilos
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        
        # Encabezado
        ws['A1'] = context['company'].name
        ws['A1'].font = header_font
        ws['A2'] = context['title']
        ws['A2'].font = title_font
        ws['A3'] = f"Del {context['period_start']} al {context['period_end']}"
        
        # Cabeceras de columnas
        row = 5
        headers = ['Código', 'Cuenta', 'Saldo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
        
        # Datos
        row += 1
        for account_data in context['accounts_data']:
            ws.cell(row=row, column=1, value=account_data[0])  # código
            ws.cell(row=row, column=2, value=account_data[1])  # nombre
            ws.cell(row=row, column=3, value=float(account_data[4]))  # saldo
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _generate_excel_trial_balance(self, context):
        """
        Genera Balance de Prueba en Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance de Prueba"
        
        # Estilos
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        
        # Encabezado
        ws['A1'] = context['company'].name
        ws['A1'].font = header_font
        ws['A2'] = context['title']
        ws['A2'].font = title_font
        ws['A3'] = f"Al {context['period_end']}"
        
        # Cabeceras de columnas
        row = 5
        headers = ['Código', 'Cuenta', 'Débito', 'Crédito', 'Saldo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
        
        # Datos
        row += 1
        total_debit = 0
        total_credit = 0
        total_balance = 0
        
        for account_data in context['accounts_data']:
            ws.cell(row=row, column=1, value=account_data[0])  # código
            ws.cell(row=row, column=2, value=account_data[1])  # nombre
            ws.cell(row=row, column=3, value=float(account_data[2]))  # débito
            ws.cell(row=row, column=4, value=float(account_data[3]))  # crédito
            ws.cell(row=row, column=5, value=float(account_data[4]))  # saldo
            
            total_debit += float(account_data[2])
            total_credit += float(account_data[3])
            total_balance += float(account_data[4])
            row += 1
        
        # Totales
        row += 1
        ws.cell(row=row, column=1, value="TOTALES").font = bold_font
        ws.cell(row=row, column=3, value=total_debit).font = bold_font
        ws.cell(row=row, column=4, value=total_credit).font = bold_font
        ws.cell(row=row, column=5, value=total_balance).font = bold_font
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _generate_excel_general_ledger(self, context):
        """
        Genera Libro Mayor en Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libro Mayor"
        
        # Estilos
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        
        # Encabezado
        ws['A1'] = context['company'].name
        ws['A1'].font = header_font
        ws['A2'] = context['title']
        ws['A2'].font = title_font
        ws['A3'] = f"Del {context['period_start']} al {context['period_end']}"
        
        # Cabeceras de columnas
        row = 5
        headers = ['Fecha', 'Comprobante', 'Concepto', 'Débito', 'Crédito', 'Saldo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
        
        # Datos
        row += 1
        running_balance = 0
        
        for movement in context['movements']:
            ws.cell(row=row, column=1, value=movement.journal_entry.date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=movement.journal_entry.number)
            ws.cell(row=row, column=3, value=movement.description or movement.journal_entry.description)
            ws.cell(row=row, column=4, value=float(movement.debit) if movement.debit > 0 else 0)
            ws.cell(row=row, column=5, value=float(movement.credit) if movement.credit > 0 else 0)
            
            running_balance += float(movement.debit - movement.credit)
            ws.cell(row=row, column=6, value=running_balance)
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _generate_excel_aging_report(self, context):
        """
        Genera Reporte de Cartera en Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cartera Vencida"
        
        # Estilos
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        
        # Encabezado
        ws['A1'] = context['company'].name
        ws['A1'].font = header_font
        ws['A2'] = context['title']
        ws['A2'].font = title_font
        ws['A3'] = f"Al {context['report_date']}"
        
        # Cabeceras de columnas
        row = 5
        headers = ['Cliente/Proveedor', 'Factura', 'Fecha', 'Vencimiento', 'Total', 'Saldo', 'Días Vencidos']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
        
        # Datos
        row += 1
        total_balance = 0
        
        for item in context['aging_data']:
            ws.cell(row=row, column=1, value=item[1])  # nombre cliente/proveedor
            ws.cell(row=row, column=2, value=item[2])  # número factura
            ws.cell(row=row, column=3, value=item[3].strftime('%d/%m/%Y') if item[3] else '')  # fecha
            ws.cell(row=row, column=4, value=item[4].strftime('%d/%m/%Y') if item[4] else '')  # vencimiento
            ws.cell(row=row, column=5, value=float(item[5]))  # total
            ws.cell(row=row, column=6, value=float(item[6]))  # saldo
            
            # Calcular días vencidos
            days_overdue = 0
            if item[4] and context['report_date']:
                days_overdue = max(0, (context['report_date'] - item[4]).days)
            ws.cell(row=row, column=7, value=days_overdue)
            
            total_balance += float(item[6])
            row += 1
        
        # Total
        row += 1
        ws.cell(row=row, column=5, value="TOTAL").font = bold_font
        ws.cell(row=row, column=6, value=total_balance).font = bold_font
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
